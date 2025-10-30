from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook

try:
    from .openai_client import AIClient, JSON_KEYS
except Exception:  # pragma: no cover - allow running as script as well
    from openai_client import AIClient, JSON_KEYS  # type: ignore


AI_COLUMNS = JSON_KEYS


def _row_to_task(row: pd.Series) -> Dict[str, Any]:
    # Ensure all keys exist and are strings
    fields = [
        "No.",
        "課題",
        "ステータス",
        "人員",
        "担当者",
        "現状把握",
        "備考・コメント",
    ]
    task: Dict[str, Any] = {}
    for k in fields:
        v = row.get(k, "")
        if pd.isna(v):
            v = ""
        task[k] = str(v)
    return task


def _apply_result_to_df(df: pd.DataFrame, idx: int, result: Dict[str, Any]) -> None:
    """Deprecated: kept for backward compatibility (no longer used for writing)."""
    for col in AI_COLUMNS:
        val = result.get(col, "")
        if col == "AI_Caution" and isinstance(val, list):
            df.at[idx, col] = "\n".join(f"・{item}" for item in val)
        else:
            df.at[idx, col] = val


def _unique_output_path(base_path: Path) -> Path:
    """If base_path exists, append a numeric suffix to avoid overwrite.
    E.g., report.xlsx -> report (2).xlsx etc.
    """
    if not base_path.exists():
        return base_path
    stem = base_path.stem
    suffix = base_path.suffix
    parent = base_path.parent
    n = 2
    while True:
        cand = parent / f"{stem} ({n}){suffix}"
        if not cand.exists():
            return cand
        n += 1


def _write_results_with_openpyxl(
    input_path: Path,
    output_path: Path,
    results_by_row: Dict[int, Dict[str, Any]],
    model: str,
    created_at: str,
) -> None:
    """Load original workbook and write only AI_* columns, preserving existing formatting."""
    wb = load_workbook(input_path)
    ws = wb.active

    # Build header map: name -> column index
    header_name_to_col: Dict[str, int] = {}
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        v = ws.cell(row=1, column=col).value
        name = str(v) if v is not None else ""
        if name:
            header_name_to_col[name] = col

    # Ensure AI columns exist; append if missing
    next_col = ws.max_column + 1
    for col_name in AI_COLUMNS:
        if col_name not in header_name_to_col:
            ws.cell(row=1, column=next_col).value = col_name
            header_name_to_col[col_name] = next_col
            next_col += 1

    # Write values (append mode: keep existing, append date/model header + body)
    for excel_row, result in results_by_row.items():
        for col_name in AI_COLUMNS:
            col_idx = header_name_to_col[col_name]
            val = result.get(col_name, "")
            if col_name == "AI_Caution" and isinstance(val, list):
                text = "\n".join(f"・{item}" for item in val)
            else:
                text = val
            c = ws.cell(row=excel_row, column=col_idx)
            old = str(c.value) if c.value is not None else ""
            header = f"{created_at} model:{model}"
            block = f"{header}\n{text}" if str(text).strip() else header
            if old.strip():
                # Separate previous content and new block by a blank line
                new_val = f"{old}\n\n{block}"
            else:
                new_val = block
            c.value = new_val

    wb.save(output_path)


def enrich_excel(
    input_path: Path,
    output_path: Path,
    json_path: Path | None,
    model: str,
    dry_run: bool = False,
    write_mode: str = "new",
    only_nos: List[str] | None = None,
    only_rows: List[int] | None = None,
    include_completed: bool = False,
) -> Dict[str, Any]:
    df = pd.read_excel(input_path, dtype=str)
    df = df.fillna("")

    results: List[Dict[str, Any]] = []

    client = None if dry_run else AIClient(model=model)

    # Map: excel_row_number -> result dict
    results_by_row: Dict[int, Dict[str, Any]] = {}

    for pos, (_, row) in tqdm(enumerate(df.iterrows()), total=len(df), desc="Evaluating"):
        status = str(row.get("ステータス", "")).strip()
        if (not include_completed) and status == "完了":
            continue
        excel_row = pos + 2

        # Apply filters if provided
        if only_nos is not None:
            no_val = str(row.get("No.", "")).strip()
            if no_val not in set(only_nos):
                continue
        if only_rows is not None:
            if excel_row not in set(only_rows):
                continue

        task = _row_to_task(row)
        if dry_run:
            result = {
                "AI_工数": "0.5人日",
                "AI_Level": "中",
                "AI_NextAction": "次の具体的な一手を簡潔に記載（DRY-RUN）",
                "AI_Advice": "進め方の注意点を簡潔に（DRY-RUN）",
                "AI_Evaluation": "現状から見た妥当な評価（DRY-RUN）",
                "AI_Caution": ["依存関係の確認", "レビュー観点の事前共有"],
            }
        else:
            result = client.evaluate_task(task)  # type: ignore[union-attr]

        # Map to Excel row number (already computed)
        results_by_row[excel_row] = result
        # For JSON export, include No. for traceability
        results.append({"No.": task.get("No.", ""), **result})

    # Write outputs using openpyxl to preserve formatting
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if write_mode == "new":
        output_path = _unique_output_path(output_path)
    # Prepare metadata
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _write_results_with_openpyxl(input_path, output_path, results_by_row, model=model, created_at=created_at)

    if json_path:
        json_path.parent.mkdir(parents=True, exist_ok=True)
        with json_path.open("w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)

    return {"rows_processed": len(results), "output": str(output_path), "json": str(json_path) if json_path else None}


def main() -> None:
    parser = argparse.ArgumentParser(description="課題管理ExcelをAIで評価し、AI_*列を追記します。")
    parser.add_argument("--input", "-i", type=Path, default=Path("kadai_kanri_simple.xlsx"), help="入力Excelファイルパス（.xlsx）")
    parser.add_argument("--output", "-o", type=Path, default=None, help="出力Excelファイルパス（省略時は *_with_ai.xlsx）")
    parser.add_argument("--json", type=Path, default=Path("ai_results.json"), help="行ごとのAI結果JSON出力パス（不要なら --json ''）")
    parser.add_argument("--model", type=str, default="gpt-5", help="OpenAIモデル名")
    parser.add_argument("--dry-run", action="store_true", help="APIを呼び出さずにダミー値で動作確認する")
    parser.add_argument(
        "--write-mode",
        choices=["new", "overwrite"],
        default="new",
        help="出力方法: new=同名時は連番で新規作成, overwrite=指定出力に上書き。--output未指定かつoverwriteの場合は入力ファイルに上書き",
    )
    parser.add_argument(
        "--only-no",
        type=str,
        default="",
        help="処理対象にするNo.をカンマ区切りで指定 (例: 1,3,5)。未指定で全行(完了は除外)。",
    )
    parser.add_argument(
        "--only-rows",
        type=str,
        default="",
        help="処理対象にするExcel行番号をカンマ/ハイフンで指定 (例: 2,4-6)。ヘッダは1行目。",
    )
    parser.add_argument(
        "--include-completed",
        action="store_true",
        help="ステータス=完了の行も処理対象に含める",
    )

    args = parser.parse_args()

    output_path = args.output
    if args.write_mode == "overwrite" and output_path is None:
        # Overwrite input when overwrite mode and no explicit output path
        output_path = Path(args.input)
    if output_path is None:
        p = Path(args.input)
        output_path = p.with_name(p.stem + "_with_ai.xlsx")

    json_path = None if str(args.json) == "" else args.json

    # Parse filters
    def _parse_no_list(s: str) -> List[str] | None:
        s = (s or "").strip()
        if not s:
            return None
        return [x.strip() for x in s.split(",") if x.strip()]

    def _parse_row_spec(s: str) -> List[int] | None:
        s = (s or "").strip()
        if not s:
            return None
        res: list[int] = []
        for part in s.split(","):
            part = part.strip()
            if not part:
                continue
            if "-" in part:
                try:
                    a, b = part.split("-", 1)
                    start = int(a)
                    end = int(b)
                    if start <= end:
                        res.extend(range(start, end + 1))
                    else:
                        res.extend(range(end, start + 1))
                except Exception:
                    continue
            else:
                try:
                    res.append(int(part))
                except Exception:
                    continue
        return sorted(set(res)) or None

    only_nos = _parse_no_list(args.only_no)
    only_rows = _parse_row_spec(args.only_rows)

    info = enrich_excel(
        input_path=args.input,
        output_path=output_path,
        json_path=json_path,
        model=args.model,
        dry_run=args.dry_run,
        write_mode=args.write_mode,
        only_nos=only_nos,
        only_rows=only_rows,
        include_completed=args.include_completed,
    )

    print(json.dumps(info, ensure_ascii=False))


if __name__ == "__main__":
    main()
